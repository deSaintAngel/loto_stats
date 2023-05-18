from collections import defaultdict
import os
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
import csv
from tabulate import tabulate
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook
import itertools
from scipy.stats import norm
from mpl_toolkits.mplot3d import Axes3D
import  matplotlib

import tensorflow as tf,keras
from tensorflow.keras.models import load_model
from tensorflow.keras.callbacks import ReduceLROnPlateau, EarlyStopping, ModelCheckpoint
from keras.layers import LSTM
from tensorflow.keras.layers import Input, Flatten, Reshape, Dense
from tensorflow.keras.models import Model
from tensorflow.keras.callbacks import Callback
import numpy as np
import tensorflow.keras.backend as K
from tensorflow.keras.models import Model
from tensorflow.keras.layers import Input, LSTM, Dense, Reshape
from tensorflow.keras.optimizers import Adam
from tensorflow.keras.callbacks import ModelCheckpoint, ReduceLROnPlateau, EarlyStopping
import os
os.environ['TF_CPP_MIN_LOG_LEVEL'] = '2'
import  matplotlib
matplotlib.use('QtAgg')

#%% fonction recherches

def find_row_boules(matrix, row):
    M = []
    index=[]
    for i in range(len(matrix)):
        if set(row).issubset(set(matrix[i])):
            index.append(i)

    if index == []:
        print(f'La ligne existe PAS nbr : 0, index : {row}')
        if len(row)==2:
            M.append([0, row[0], row[1]])
        if len(row)==3:
            M.append([0, row[0], row[1], row[2]])

    else:
        # print(index)
        # print("La ligne existe dans la matrice, son index est", index)
        print(f'La ligne{row} existe nbr : {len(index)}, index : {index}')
        if len(row) == 2:
            M.append([len(index),row[0],row[1]])
        if len(row) == 3:
            M.append([len(index),row[0],row[1],row[2]])

    return M

def find_row_etoiles(matrix, row):
    M = []
    index=[]
    for i in range(len(matrix)):
        if all(matrix[i][j] == row[j] for j in range(len(row))):
            index.append(i)

    if index == []:
        print("La ligne n'existe pas dans la matrice")

    else:
        # print(index)
        # print("La ligne existe dans la matrice, son index est", index)
        print(f'La ligne existe nbr : {len(index)}, index : {row}')
        M.append([len(index),row[0],row[1]])

    return M



def trouver_lignes_identiques(matrice):
    for i in range(len(matrice)):
        for j in range(i+1, len(matrice)):
            if np.array_equal(matrice[i], matrice[j]):
                indice_i, indice_j, ligne = i, j, matrice[i]
                print(f"Les lignes {indice_i+1} et {indice_j+1} sont identiques : {ligne}")
                return indice_i, indice_j, ligne
    print("Il n'y a pas de lignes identiques dans cette matrice.")
    return None


def generer_binomes(sequence):
    binomes = []

    for i in range(len(sequence)):
        for j in range(i + 1, len(sequence)):
            binome = [sequence[i], sequence[j]]
            binomes.append(binome)

    return binomes


def generer_trinomes(sequence):
    trinomes = []

    for i in range(len(sequence)):
        for j in range(i + 1, len(sequence)):
            for k in range(j + 1, len(sequence)):
                trinome = [sequence[i], sequence[j], sequence[k]]
                trinomes.append(trinome)

    return trinomes


def generer_quadrinomes(sequence):
    quadrinomes = []

    for i in range(len(sequence)):
        for j in range(i + 1, len(sequence)):
            for k in range(j + 1, len(sequence)):
                for l in range(k + 1, len(sequence)):
                    quadrinome = [sequence[i], sequence[j], sequence[k], sequence[l]]
                    quadrinomes.append(quadrinome)

    return quadrinomes

def generer_tableau_numpy(tab):
    n = len(tab)
    m = len(tab[0])
    tableau_numpy = np.array(tab).reshape((n, m))
    return tableau_numpy

#%% fonction statistiques

def hist_par_colonnes(tab,xn=1,yn=2):
    for i in range(tab.shape[1]):
        d = np.abs(np.max(tab[:, i])-np.min(tab[:, i])).astype(int)
        fig, ax = plt.subplots()
        occurrences, bins, _ = ax.hist(tab[:, i], bins=2*d+1)
        ax.set_xlabel('Valeur')
        ax.set_ylabel('Nombre d\'occurrences')
        ax.set_title('Histogramme')
        ax.grid(axis='y')
        ax.yaxis.set_major_locator(plt.MultipleLocator(yn))
        ax.xaxis.set_major_locator(plt.MultipleLocator(xn))
        valeur_max = bins[np.argmax(occurrences)]
        print(f'La valeur correspondant à l\'occurrence maximale de la colonne {i + 1} est {valeur_max}')
        plt.show()

def hist_(tab,xn=2,yn=1):
    d = np.abs(np.max(tab)-np.min(tab)).astype(int)
    fig, ax = plt.subplots()
    occurrences, bins, _ = ax.hist(tab.ravel(), bins=2*d+1)
    ax.set_xlabel('Valeur')
    ax.set_ylabel('Nombre d\'occurrences')
    ax.set_title('Histogramme')
    ax.grid(axis='y')
    ax.yaxis.set_major_locator(plt.MultipleLocator(yn))
    ax.xaxis.set_major_locator(plt.MultipleLocator(xn))
    valeur_max = bins[np.argmax(occurrences)]
    plt.show()

def kl_divergence(p, q):
    return np.sum(np.where(p != 0, p * np.log(p / q), 0))

def normal_approximation(data):
    mean, std = norm.fit(data)
    d = np.abs(np.max(data) - np.min(data)).astype(int)
    fig, ax = plt.subplots()
    x = np.linspace(data.min(), data.max(), len(data))
    ax.hist(data, density=True, bins=2 * d + 1)
    ax.hist(data, bins='auto', density=True, alpha=0.7)
    ax.plot(x, norm.pdf(x, mean, std), 'r-', lw=1, label='normal pdf')
    ax.axvline(x=round(mean), color='k', linestyle='--', label=f'mean ={mean:.2f} soit {round(mean):.0f}')
    # ax.yaxis.set_major_locator(plt.MultipleLocator(2))
    # ax.xaxis.set_major_locator(plt.MultipleLocator(1))
    ax.legend(loc='best')
    ax.grid(axis='y')
    plt.show()
    return mean, std


def normal_parameters(matrix):
    n, m = matrix.shape
    parameters = []
    for j in range(m):
        data = matrix[:, j]
        mean, std = normal_approximation(data)
        parameters.append((mean, std))
    return parameters


def kl_distrib(matrix):
    parameters = normal_parameters(matrix)
    for j in range(np.shape(matrix)[-1]):
        p = norm.pdf(matrix[:, j], *parameters[j])
        q = norm.pdf(matrix[:, j], matrix[:, j].mean(), matrix[:, j].std())
        kl = kl_divergence(p, q)
        print(f"Column {j}: Mean = {parameters[j][0]}, Std = {parameters[j][1]}, KL-Divergence = {kl}")


def calc_tab_save(data,entetes,title):
    """
    recupere un tableau de données et enregistre un fichier excel avec les données et les entetes,
    marque en rouge les valeurs max de chaque colonne
    :param data: tableau de données
    :param entetes: entêtes du tableau de données
    :param title: titre du fichier Excel
    :return: None, fichier Excel enregistré
    """
    df = pd.DataFrame(data, columns=entetes)
    max_indexes = df.idxmax()
    # Créer un nouveau classeur Excel
    wb = Workbook()
    # Sélectionner la première feuille du classeur
    ws = wb.active
    # Ajouter les données du dataframe à la feuille
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)
    # Créer un objet Fill avec une couleur rouge
    fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
    # Parcourir les colonnes de la deuxième jusqu'à la dernière colonne
    # Parcourir les colonnes de la deuxième jusqu'à la dernière colonne
    for i, col in enumerate(df.columns[1:], 2):
        # Trouver les index de toutes les occurrences de la valeur maximale dans la colonne
        max_index_list = [index+1 for index, value in enumerate(df[col]) if value == df[col].max()]
        # Appliquer le remplissage rouge à chaque cellule correspondante
        for row in max_index_list:
            ws.cell(row=row+1, column=i).fill = fill
    # Enregistrer le classeur Excel
    wb.save(f'{title}.xlsx')


#%% fonctions de combinaisons

def verif_comb_etoiles(jeux):
    if np.shape(jeux)[1] != 2:
        print('Le tableau doit avoir 2 colonnes')
    else:
        combinaison = np.zeros((66, 7))
        indx = 0
        for i in range(13):
            for j in range(i - 1, 0, -1):
                combinaison[indx, -2] = i
                combinaison[indx, -1] = j
                indx += 1
        etoiles_tier = np.apply_along_axis(np.sort, axis=1, arr=combinaison[:, 5:]).astype(int)
        comb_jeux_etoiles = np.apply_along_axis(np.sort, axis=1, arr=jeux).astype(int)

        etoiles_stat = []
        for i in range(len(combinaison)):
            etoiles_stat.append(find_row_etoiles(comb_jeux_etoiles,etoiles_tier[i]))

        etoiles_stat = np.reshape(etoiles_stat,(np.shape(etoiles_stat)[0],np.shape(etoiles_stat)[-1]))
        np.savetxt(f'./data/binome_etoiles_.csv', etoiles_stat, delimiter=',', fmt='%d', comments='')
        plot_3d_etoiles(etoiles_stat)

def verif_comb_duo_boules(jeux,loto_type):
    if loto_type == 'euromillions':
        combinaison = np.zeros((1225, 2))
    elif loto_type == 'loto':
        combinaison = np.zeros((1176, 2))
    indx = 0
    for i in range(np.max(jeux).astype(int) +1):
        for j in range(i - 1, 0, -1):
            combinaison[indx, -2] = i
            combinaison[indx, -1] = j
            indx += 1
    boules_tier = np.apply_along_axis(np.sort, axis=1, arr=combinaison[:,:]).astype(int)
    comb_jeux_boules = np.apply_along_axis(np.sort, axis=1, arr=jeux[:,:5]).astype(int)

    boules_stat = []
    for i in range(len(combinaison)):
        boules_stat.append(find_row_boules(comb_jeux_boules, boules_tier[i]))

    boules_stat = np.reshape(boules_stat,(len(boules_stat),3))
    np.savetxt(f'./data/binome_{loto_type}_.csv', boules_stat, delimiter=',', fmt='%d', comments='')
    plot_3d_etoiles(boules_stat)

def verif_comb_trio_boules(jeux):
    combinaison = np.array(list(itertools.combinations(range(1, 51), 3)))
    boules_tier = np.apply_along_axis(np.sort, axis=1, arr=combinaison).astype(int)
    comb_jeux_boules = np.apply_along_axis(np.sort, axis=1, arr=jeux[:,:5]).astype(int)

    boules_stat = []
    for i in range(len(combinaison)):
        boules_stat.append(find_row_boules(comb_jeux_boules, boules_tier[i]))

    boules_stat = np.reshape(boules_stat,(len(boules_stat),4))
    np.savetxt('nome_boules_euromillion.csv', boules_stat, delimiter=',', fmt='%d', comments='')


#%% fonctions historique

def derniere_ligne(tableau,loto_type='loto'):
    # Définition du nombre de lignes et du nombre de nombres dans chaque ligne
    n = len(tableau)
    m = len(tableau[0])

    # Initialisation d'un dictionnaire pour stocker les dernières lignes d'apparition de chaque nombre
    dernieres_lignes = defaultdict(int)

    # Boucle pour traiter chaque ligne
    for i in range(n):
        ligne = tableau[i]

        # Boucle pour traiter chaque nombre dans la ligne
        for j in range(m):
            nombre = ligne[j]

            # Si le nombre a déjà été vu, mettre à jour la dernière ligne d'apparition
            if nombre in dernieres_lignes:
                dernieres_lignes[nombre] = i + 1

            # Sinon, ajouter le nombre avec la ligne actuelle comme dernière ligne d'apparition
            else:
                dernieres_lignes[nombre] = i + 1

    # Initialisation d'un tableau numpy pour stocker les résultats
    resultats = np.zeros((50, 2))


    # Boucle pour remplir le tableau numpy avec les résultats

    for k in range(1, 51):
        resultats[k - 1, 0] = k
        resultats[k - 1, 1] = n - dernieres_lignes[k]

    d = np.max(tableau).astype(int)
    return resultats[:np.max(tableau).astype(int)]


#%% fonctions suite binome

def suite_binome(boules):
    """
    fonction qui recuperer les binomes de boules qui se suivent (exemple : {1,2},{2,3},{3,4},...) et qui donne le nombre de fois que cela se produit
    :param boules_tier: numpy array de taille (n,7) avec n le nombre de combinaisons totales
    :return: numpy array de taille (n,3) avec n le nombre de combinaisons totales, premiere et deuxieme colonne : binome de boules, dernierer colonne : nombre de fois que cela se produit
    """
    boules_tier = np.apply_along_axis(np.sort, axis=1, arr=boules[:,:5]).astype(int)

     # creation de l'ensemble de binomes de boules dont les numeros se suivent (exemple : {1,2},{2,3},{3,4},...) pour les boules de 1 a 50 (exclu)
    binome_boules = []
    for i in range(1,np.max(boules_tier).astype(int)):
        binome_boules.append([i,i+1])
    binome_boules = np.array(binome_boules)

    # creation d'un tableau de taille (n,3) avec n le nombre de combinaisons totales,
    # premiere et deuxieme colonne : binome de boules, dernierer colonne : nombre de fois que cela se produit
    binome_boules_stat = np.zeros((len(binome_boules),3))
    # pour chaque binome de boules on compte le nombre de fois qu'il apparait dans les combinaisons
    for i, binome in enumerate(binome_boules):
        binome_boules_stat[i, 0] = binome[0]
        binome_boules_stat[i, 1] = binome[1]
        nb = find_row_boules(boules_tier, binome)[0][0]
        binome_boules_stat[i, 2] = nb

    return binome_boules_stat

def suite_trinome(boules):
    """
    fonction qui recuperer les trinomes de boules qui se suivent (exemple : {1,2,3},{2,3,4},{3,4,5},...) et qui donne le nombre de fois que cela se produit
    :param boules_tier: numpy array de taille (n,7) avec n le nombre de combinaisons totales
    :param boules: numpy array de taille (n,7) avec n le nombre de combinaisons totales,
    :return: numpy array de taille (n,4) avec n le nombre de combinaisons totales,
    premiere et deuxieme colonne : binome de boules, dernierer colonne : nombre de fois que cela se produit
    """
    boules_tier = np.apply_along_axis(np.sort, axis=1, arr=boules[:,:5]).astype(int)

    # creation de l'ensemble de binomes de boules dont les numeros se suivent (exemple : {1,2},{2,3},{3,4},...) pour les boules de 1 a 50 (exclu)
    trinome_boules = []
    for i in range(1,np.max(boules_tier).astype(int)-1):
        trinome_boules.append([i,i+1,i+2])
    trinome_boules = np.array(trinome_boules)

    # creation d'un tableau de taille (n,3) avec n le nombre de combinaisons totales,
    # premiere et deuxieme colonne : binome de boules, dernierer colonne : nombre de fois que cela se produit
    trinome_boules_stat = np.zeros((len(trinome_boules),4))
    # pour chaque binome de boules on compte le nombre de fois qu'il apparait dans les combinaisons
    for i, trinome in enumerate(trinome_boules):
        trinome_boules_stat[i, 0] = trinome[0]
        trinome_boules_stat[i, 1] = trinome[1]
        trinome_boules_stat[i, 2] = trinome[2]
        nb = find_row_boules(boules_tier, trinome)[0][0]
        trinome_boules_stat[i, 3] = nb

    return trinome_boules_stat

#%% fonctions calcules ecart entre tirages
def generate_table(liste):
    table = [[0] for _ in range(50)]
    for i in range(len(liste)):
        for j in range(51):
            if j in liste[i]:
                table[j-1].append(i+1)
    return table



def calculate_differences(table):
    td = np.zeros((50, 120))
    differences = []
    for i in range(len(table)):
        diff = [int(table[i][j+1] - table[i][j]) for j in range(len(table[i])-1)]
        differences.append(diff)
        td[i, :len(diff)] = diff
    return td

def diff_loto(t, b):
    tab = generate_table(t)
    d = calculate_differences(tab)
    d_array = np.array(d)
    moyenne_ecart = np.mean(d_array,axis=1)
    ecarttype = np.std(d_array,axis=1)
    np.savetxt(f'diff_loto{str(b)}.csv', d_array, delimiter=",", fmt='%d', comments='')

    # contanetation de ecarttype et de l'indice de ligne
    met = np.concatenate((np.arange(1,51).reshape(50,1),np.round(ecarttype).reshape(50,1)),axis=1)
    np.savetxt(f'ecarttype_loto{b}.csv', met, delimiter=",", fmt='%d', comments='')

    print("Numéro\tÉcart-type")
    for i in range(len(ecarttype)):
        print(f"{i+1}\t{round(ecarttype[i], 0)}")

    return d


#%% fonctions graphiques
def plot_3d_etoiles(matrice):
    # Création de la figure 3D
    fig = plt.figure()
    ax = fig.add_subplot(111, projection='3d')
    # Récupération des colonnes de la matrice
    valeurs = [row[0] for row in matrice]
    var1 = [row[1] for row in matrice]
    var2 = [row[2] for row in matrice]
    # Affichage des points en 3D
    ax.scatter(var1, var2, valeurs)
    # Ajout de labels pour les axes
    ax.set_xlabel('Variable x')
    ax.set_ylabel('Variable y')
    ax.set_zlabel('Valeur')
    # Affichage de la figure
    # plt.show()
    ax.set_xticks(np.arange(min(var1), max(var1) + 1, 2))
    ax.set_yticks(np.arange(min(var2), max(var2) + 1, 2))
    ax.set_zticks(np.arange(min(valeurs), max(valeurs) + 1, 1))

    plt.show()

#%% fonction stat to xlsx
def calc_tab_save(data,entetes,title):
    df = pd.DataFrame(data, columns=entetes)
    max_indexes = df.idxmax()
    # Créer un nouveau classeur Excel
    wb = Workbook()
    # Sélectionner la première feuille du classeur
    ws = wb.active
    # Ajouter les données du dataframe à la feuille
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)
    # Créer un objet Fill avec une couleur rouge
    fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
    # Parcourir les colonnes de la deuxième jusqu'à la dernière colonne
    # Parcourir les colonnes de la deuxième jusqu'à la dernière colonne
    for i, col in enumerate(df.columns[1:], 2):
        # Trouver les index de toutes les occurrences de la valeur maximale dans la colonne
        max_index_list = [index+1 for index, value in enumerate(df[col]) if value == df[col].max()]
        # Appliquer le remplissage rouge à chaque cellule correspondante
        for row in max_index_list:
            ws.cell(row=row+1, column=i).fill = fill
    # Enregistrer le classeur Excel
    wb.save(f'{title}.xlsx')


def frequence_calc(boules, loto_type=None, trier_=True):
    # # Calcul des fréquences d'apparitions
    freq_total = []
    freq_par_boules = []

    for i in range(1):
        col_total = boules.ravel()
        col_freq_total = []
        for j in range(1, np.max(boules).astype(int)+1):
            col_freq_total.append(np.count_nonzero(col_total == j))
        freq_total.append(col_freq_total)


    for i in range(np.shape(boules)[-1]):
        col_par_boules = boules[:, i]
        col_freq_par_boules = []
        for j in range(1,  np.max(boules).astype(int)+1):
            col_freq_par_boules.append(np.count_nonzero(col_par_boules == j))
        freq_par_boules.append(col_freq_par_boules)

    # Préparation des données pour l'affichage
    entetes_tot = ["Valeur", "c"]
    if np.shape(boules)[-1] == 5:
        entetes_par_boules = ["Valeur", "c  1", " c 2", " c 3", " c 4", "c 5"]
    elif np.shape(boules)[-1] == 2:
        entetes_par_boules = ["Valeur", "c  1", " c 2"]


    data_tot = []
    data_par_boules = []

    for i in range(1, np.max(boules).astype(int)+1):
        row_tot = [i] + [freq_total[j][i-1] for j in range(1)]
        data_tot.append(row_tot)

        row_par_boules = [i] + [freq_par_boules[j][i-1] for j in range(np.shape(boules)[-1])]
        data_par_boules.append(row_par_boules)

    if trier_:
        trie = '_trier'
    else:
        trie =''

    if np.shape(boules)[-1] == 2 and loto_type == 'euromillions':
        type_ = '_etoiles'
    elif np.shape(boules)[-1] == 5 and loto_type == 'euromillions':
        type_ = ''
    elif np.shape(boules)[-1] == 1 and loto_type == 'loto':
        type_ = '_numero_chance'
    else:
        type_ = ''
    calc_tab_save(data_tot,entetes_tot,f'./data/xlsx/{loto_type}{type_}_frequence_total'+trie)
    if np.shape(boules)[-1] > 1:
        calc_tab_save(data_par_boules, entetes_par_boules, f'./data/xlsx/{loto_type}{type_}_frequence_par_boules'+trie)

    return data_tot,data_par_boules


#%% fonctions deep learning
def create_data(tab, series_size):
    tab = np.transpose(tab)
    x_train = []
    y_train = []
    num_samples = tab.shape[-1] - series_size
    for i in range(0, num_samples):
        x_train.append(tab[:, 0 + i:series_size + i].astype(np.int32))
        y_train.append(tab[:, series_size + i].astype(np.int32))

    x_train = tf.convert_to_tensor(x_train)
    y_train = tf.convert_to_tensor(np.expand_dims(y_train, axis=1))
    # y_one_hot = to_categorical(y_train-1, num_classes=50)

    return x_train, y_train

class SaveModelOnThreshold(ModelCheckpoint):
    def __init__(self, filepath, monitor='val_loss', threshold=0.0, **kwargs):
        super().__init__(filepath=filepath, monitor=monitor, **kwargs)
        self.threshold = threshold

    def on_epoch_end(self, epoch, logs=None):
        current = logs.get(self.monitor)
        if current is None:
            return

        if current < self.threshold:
            filepath = self.filepath.format(epoch=epoch, **logs)
            self.model.save(filepath, overwrite=True)
            self.best = current


from tensorflow.keras.callbacks import Callback
from tqdm import tqdm


class ProgressCallback(Callback):
    def __init__(self, num_epochs):
        super(ProgressCallback, self).__init__()
        self.num_epochs = num_epochs
        self.loss_values = []  # Ajout de l'attribut loss_values

    def on_train_begin(self, logs=None):
        self.pbar = tqdm(total=self.num_epochs, desc="Training Epochs")

    def on_epoch_end(self, epoch, logs=None):
        self.pbar.update(1)
        self.pbar.set_postfix(loss=logs['loss'], lr=self.model.optimizer.lr.numpy())
        self.loss_values.append(logs['loss'])  # Ajout de la valeur de la perte à loss_values

    def on_train_end(self, logs=None):
        self.pbar.close()

        # Plot the loss
        plt.figure(figsize=(10, 5))
        plt.plot(self.loss_values)
        plt.xlabel('Epochs')
        plt.ylabel('Loss')
        plt.title('Training Loss')
        plt.show()


